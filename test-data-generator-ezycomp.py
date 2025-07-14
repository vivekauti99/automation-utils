from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import random
from faker import Faker
from datetime import datetime, timedelta, date
import calendar
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

fake = Faker('en_IN')

# Constants
departments = ['HR', 'Finance', 'Sales', 'IT', 'Admin']
designations = ['Executive', 'Manager', 'Analyst', 'Associate']
est_types = ['Shops']
vendor_works = ['Housekeeping', 'Security', 'Maintenance']
exit_reasons = ['Resigned', 'Terminated', 'Retired']
esi_causes = ['Injury', 'Illness', 'Maternity']
esi_natures = ['Temporary', 'Permanent']
esi_places = ['Mumbai', 'Ahmedabad', 'New Delhi', 'Pune']
bank_names = ['State Bank of India', 'ICICI Bank', 'HDFC Bank', 'Axis Bank', 'Bank of Baroda', 'Canara Bank']

columns_employee = [
    'PARENT COMPANY', 'ASSOCIATE COMPANY', 'STATE', 'LOCATION', 'Branch', 'Employee Code',
    'Employee name', 'Gender', 'DOB', 'DOJ', 'DOL', 'Age', 'Father Name', 'Department',
    'Bank Name', 'Bank Account Number', 'PAN Number', 'Aadhar Number', 'PF No', 'PF UAN No',
    'ESI Number', 'Designation', 'Establishment Type', 'Vendor Nature Of Work',
    'Reason of Exit', 'Office In Time', 'Office Out Time', 'Interval In Time',
    'Interval Out Time', 'Date of Payment', 'ESI Date of Notice', 'ESI Time of Notice',
    'ESI Cause', 'ESI Nature', 'ESI Date', 'ESI Time', 'ESI Place'
]

columns_leave_credit = [
    'PARENT COMPANY', 'ASSOCIATE COMPANY', 'STATE', 'LOCATION', 'Employee code',
    'Month', 'Year', 'PL,EL,AL Monthly credit', 'PL,EL,AL Opening balance',
    'PL,EL,AL Closing balance', 'SL Monthly credit', 'SL Opening balance',
    'SL Closing balance', 'CL Monthly credit', 'CL Opening balance',
    'CL Closing balance', 'ML Opening balance', 'ML Closing balance'
]

columns_leave_availed = [
    'PARENT COMPANY', 'ASSOCIATE COMPANY', 'STATE', 'LOCATION',
    'Month', 'Year', 'EMPLOYEE CODE', 'EMPLOYEE NAME',
    'START DATE', 'END DATE', 'LEAVE TYPE', 'NO OF DAYS',
    'HALF DAY LEAVE TAKEN DATE'
]

columns_attendance = [
        'PARENT COMPANY', 'ASSOCIATE COMPANY', 'STATE', 'LOCATION',
        'Month', 'Year', 'Employee Code', 'Employee name'
] + [str(i) for i in range(1, 32)] + ['Present Days']

columns_wage = [
    'PARENT COMPANY', 'ASSOCIATE COMPANY', 'STATE', 'LOCATION',
    'Branch', 'Month', 'Year', 'Employee Code', 'Present Days', 'Work Days', 'LOP Days',
    'Fixed Gross', 'Basic Wages', 'Basic Arrear', 'House Rent Allowance', 'House Rent Arrear',
    'Conveyance Allowance', 'Conveyance Arrear', 'Special Allowance', 'Special Allowance Arrear',
    'Dearness Allowance', 'Dearness Arrear', 'Misc. Allowances', 'Other Earning', 'Bonus',
    'Over Time', 'Gratuity', 'Medical Allowance', 'Holiday Payment', 'Gross Earnings',
    'Gross Earnings Arrear', 'Provident Fund', 'Provident Fund Arrear', 'Employer Provident Fund',
    'Professional Tax', 'Profeessiona Tax Arrear', 'Income Tax', 'ESI', 'ESI Arrear', 'LWF',
    'Other Deduction', 'Fines', 'Insurance', 'Salary Advance', 'Gross Deduction',
    'Gross Deduction Arrear', 'Net Payable', 'Net Payable Arrear', 'Overtime Hours',
    'Overtime Taken Date', 'Advance Paid Date', 'No of Adv Installments to be Recovered',
    'Date on Adv Recovery Completed', 'Nature & Date for Fine', 'Relay or Set', 'Act or Omiss',
    'Date Of Fines Show Cause', 'Date of Fine Recovery Completed', 'Damage Loss Amount',
    'Cause of Damage Or Loss Date', 'Date Of Deduction Show Cause',
    'No of Deduction Installment to be recovered', 'Date on Which Deduction Completed',
    'No. Of Installment for FDA', 'First InstallMent Date for FDA', 'Last InstallMent Date for FDA',
    'Other Concessions', 'Absense from Duty', 'Tenure of employment', 'Purpose of Advance',
    'Deduction Date of First Installment', 'Deduction Date of Last Installment',
    'Particular of Damage Or Loss', 'Total Contribution Deduction', 'Maternity Leave From Date',
    'Date Of Birth Of child', 'Insurance paid date', 'Insurance Paid amount', 'No.of days Laid Off',
    'Discharge and Dismissal', 'Proof of Illness', 'Proof of Death', 'Proof of Birth',
    'Maternity paid Date', 'Maternity Advance Paid', 'Maternity Bonus'
]

emp_counter = 1001

def style_excel_header(file_path):
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            # Adjust column widths based on max length
        for column_cells in ws.columns:
            max_length = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col].width = adjusted_width
    wb.save(file_path)
    
def split_and_save_excel(df, base_filename, max_rows=450, output_dir="."):
    os.makedirs(output_dir, exist_ok=True)
    num_parts = (len(df) // max_rows) + (1 if len(df) % max_rows else 0)
    for i in range(num_parts):
        start = i * max_rows
        end = start + max_rows
        part_df = df.iloc[start:end]
        part_file = os.path.join(output_dir, f"{base_filename}_part{i+1}.xlsx")
        part_df.to_excel(part_file, index=False)
        style_excel_header(part_file)

def save_in_multiple_batch_sizes(df, base_filename, batch_sizes, base_output_dir="."):
    for batch_size in batch_sizes:
        subfolder = f"{base_filename}_split_{batch_size}"
        output_dir = os.path.join(base_output_dir, subfolder)
        split_and_save_excel(df, base_filename, max_rows=batch_size, output_dir=output_dir)


def generate_time_am_pm(hour: int, minute: int) -> str:
    return datetime.strptime(f"{hour}:{minute}", "%H:%M").strftime("%I:%M %p")

def round_half(value):
    return round(value * 2) / 2

def generate_employee_record_with_year(parent_company, associate_company, state, location, target_year):
    global emp_counter
    gender = random.choice(['MALE', 'FEMALE'])
    first_name = fake.first_name_male() if gender == 'MALE' else fake.first_name_female()
    last_name = fake.last_name()
    full_name = f"{first_name} {last_name}"
    father_name = f"{fake.first_name_male()} {last_name}"

    reference_date = datetime(target_year, 1, 1) 
    max_dob = reference_date - timedelta(days=20 * 365 + 1)
    min_dob = reference_date - timedelta(days=55 * 365)

    dob = fake.date_between(start_date=min_dob, end_date=max_dob)
    age = reference_date.year - dob.year - ((reference_date.month, reference_date.day) < (dob.month, dob.day))

    start_date = datetime(target_year, 1, 1)
    end_date = datetime(target_year, 12, 31)
    today_limit = datetime.today()
    if target_year == today_limit.year:
        end_date = today_limit
    doj = fake.date_between(start_date=start_date, end_date=end_date)

    today_date = date.today()
    if random.random() > 0.2:
        dol = "-"
        esi_end_date = min(today_date, doj + timedelta(days=1500))
    else:
        dol_date = doj + timedelta(days=random.randint(90, 600))
        dol_date = min(dol_date, today_date)
        dol = dol_date
        esi_end_date = dol_date

    emp_code = f"EMP{emp_counter}"
    emp_counter += 1

    pan_number = fake.bothify(text='?????####?').upper()
    pan_number = ''.join([c if c.isalnum() else 'A' for c in pan_number])
    pan_number = pan_number[:5] + str(random.randint(1000, 9999)) + pan_number[-1]

    account_number = str(random.randint(1, 9)) + ''.join([str(random.randint(0, 9)) for _ in range(random.randint(9, 15))])
    aadhar = str(random.randint(1, 9)) + ''.join([str(random.randint(0, 9)) for _ in range(11)])
    pf_uan = str(random.randint(1, 9)) + ''.join([str(random.randint(0, 9)) for _ in range(11)])
    
    esi_applicable = random.random() > 0.5
    esi_number = str(random.randint(1, 9)) + ''.join([str(random.randint(0, 9)) for _ in range(9)]) if esi_applicable else "-"

    aadhar = ''.join([str(random.randint(0, 9)) for _ in range(12)])
    pf_no = 'PF' + ''.join([str(random.randint(0, 9)) for _ in range(8)])
    pf_uan = ''.join([str(random.randint(0, 9)) for _ in range(12)])

    esi_data = {
        'ESI Date of Notice': fake.date_between(start_date=doj, end_date=esi_end_date).strftime('%d/%m/%Y') if esi_applicable else "-",
        'ESI Time of Notice': fake.time(pattern="%I:%M %p") if esi_applicable else "-",
        'ESI Cause': random.choice(esi_causes) if esi_applicable else "-",
        'ESI Nature': random.choice(esi_natures) if esi_applicable else "-",
        'ESI Date': fake.date_between(start_date=doj, end_date=esi_end_date).strftime('%d/%m/%Y') if esi_applicable else "-",
        'ESI Time': fake.time(pattern="%I:%M %p") if esi_applicable else "-",
        'ESI Place': random.choice(esi_places) if esi_applicable else "-"
    }

    return [
        parent_company, associate_company, state, location, location, emp_code, full_name, gender,
        dob.strftime('%d/%m/%Y'), doj.strftime('%d/%m/%Y'),
        dol.strftime('%d/%m/%Y') if dol != "-" else "-", age, father_name,
        random.choice(departments), random.choice(bank_names), account_number,
        pan_number, aadhar, pf_no, pf_uan, esi_number,
        random.choice(designations), random.choice(est_types), random.choice(vendor_works),
        random.choice(exit_reasons) if dol != "-" else "-",
        generate_time_am_pm(9, 0), generate_time_am_pm(18, 0),
        generate_time_am_pm(13, 0), generate_time_am_pm(13, 30),
        fake.date_between(start_date=doj, end_date=esi_end_date).strftime('%d/%m/%Y') if dol != "-" else "-",
        esi_data['ESI Date of Notice'], esi_data['ESI Time of Notice'],
        esi_data['ESI Cause'], esi_data['ESI Nature'], esi_data['ESI Date'],
        esi_data['ESI Time'], esi_data['ESI Place']
    ]

def generate_employee_dataset(input_data, target_year, employees_per_combo=15):
    all_data = []
    parent_company = input_data["PARENT COMPANY"]
    associate_companies = [x.strip() for x in input_data["ASSOCIATE COMPANY"].split(",")]
    state_location_pairs = [x.strip().split("|") for x in input_data["STATE_LOCATION"].split(",")]

    for associate_company in associate_companies:
        for state, location in state_location_pairs:
            for _ in range(employees_per_combo):
                record = generate_employee_record_with_year(parent_company, associate_company, state, location, target_year)
                all_data.append(record)

    return pd.DataFrame(all_data, columns=columns_employee)

def generate_leave_credit_dataset_realistic(employee_df, month: str, year: int):
    leave_data = []
    month_num = list(calendar.month_name).index(month)
    for _, row in employee_df.iterrows():
        pl_credit = round_half(random.choice([0.5, 1.0, 1.5]))
        pl_open = round_half(random.uniform(5.0, 20.0))
        pl_close = round_half(pl_open + pl_credit)

        sl_credit = 0.5
        sl_open = round_half(random.uniform(2.0, 10.0))
        sl_close = round_half(sl_open + sl_credit)

        cl_credit = round_half(random.choice([0.5, 1.0]))
        cl_open = round_half(random.uniform(1.0, 6.0))
        cl_close = round_half(cl_open + cl_credit)

        ml_open = round_half(random.uniform(30, 90)) if row['Gender'] == "FEMALE" and random.random() < 0.2 else 0
        ml_close = ml_open

        entry = [
            row['PARENT COMPANY'], row['ASSOCIATE COMPANY'], row['STATE'], row['LOCATION'], row['Employee Code'],
            month, year,
            pl_credit, pl_open, pl_close,
            sl_credit, sl_open, sl_close,
            cl_credit, cl_open, cl_close,
            ml_open, ml_close
        ]
        leave_data.append(entry)

    return pd.DataFrame(leave_data, columns=columns_leave_credit)

def generate_leave_availed_data(employee_df, month: str, year: int):
    leave_data = []
    leave_types = ['PL', 'SL', 'CL', 'ML', 'HFD', 'LOP']
    month_num = list(calendar.month_name).index(month)
    start_date = date(year, month_num, 1)
    days_in_month = calendar.monthrange(year, month_num)[1]

    used_employees = employee_df.sample(frac=0.30, random_state=42)
    for _, row in used_employees.iterrows():
        leave_days = []
        taken_days = sorted(random.sample(range(1, days_in_month + 1), k=random.randint(1, 5)))
        for day in taken_days:
            leave_days.append(date(year, month_num, day))

        intervals = []
        temp = [leave_days[0]]
        for i in range(1, len(leave_days)):
            if (leave_days[i] - leave_days[i-1]).days == 1:
                temp.append(leave_days[i])
            else:
                intervals.append(temp)
                temp = [leave_days[i]]
        intervals.append(temp)

        for segment in intervals:
            leave_type = random.choice(leave_types if row['Gender'] == "FEMALE" else ['PL', 'SL', 'CL', 'HFD', 'LOP'])
            if leave_type == 'HFD':
                no_of_days = 0.5
                hfd_date = segment[0]
                start_dt = end_dt = segment[0]
            else:
                no_of_days = round_half(len(segment))
                hfd_date = segment[0] if no_of_days == 0.5 else pd.NaT
                start_dt, end_dt = segment[0], segment[-1]

            leave_data.append([
                row['PARENT COMPANY'], row['ASSOCIATE COMPANY'], row['STATE'], row['LOCATION'],
                month, year, row['Employee Code'], row['Employee name'],
                start_dt, end_dt, leave_type, no_of_days, hfd_date
            ])

    return pd.DataFrame(leave_data, columns=columns_leave_availed)

def format_date_columns(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%d/%m/%Y')
    return df

def get_leave_mapping(leave_availed_df):
    leave_map = {}
    for _, row in leave_availed_df.iterrows():
        emp_code = row['EMPLOYEE CODE']
        leave_type = row['LEAVE TYPE']
        start_date = pd.to_datetime(row['START DATE'], dayfirst=True)
        end_date = pd.to_datetime(row['END DATE'], dayfirst=True)
        half_day = pd.to_datetime(row['HALF DAY LEAVE TAKEN DATE'], dayfirst=True) if pd.notna(row['HALF DAY LEAVE TAKEN DATE']) else None

        current = start_date
        while current <= end_date:
            key = (emp_code, current.day)
            leave_map[key] = leave_type if leave_type != 'HFD' else 'HFD'
            current += timedelta(days=1)

        if half_day:
            leave_map[(emp_code, half_day.day)] = 'HFD'
    return leave_map

# Define week off using day names
WEEKOFF_DAY_NAMES = ['Saturday', 'Sunday']
WEEKDAY_NAME_TO_INDEX = {day: i for i, day in enumerate(calendar.day_name)}
WEEKOFF_DAYS = [WEEKDAY_NAME_TO_INDEX[d] for d in WEEKOFF_DAY_NAMES]

def get_random_nh_sh(year, month, days_in_month, states, weekoff_days):
    all_days = [date(year, month, d) for d in range(1, days_in_month + 1)]
    working_days = [d for d in all_days if d.weekday() not in weekoff_days]

    national_holidays = random.sample(working_days, k=min(2, len(working_days)))
    state_holidays = {state: random.sample(working_days, k=random.randint(0, 1)) for state in states}
    return national_holidays, state_holidays

def get_random_date_or_dash(year, month):
    return "{:02d}/{:02d}/{}".format(random.randint(1, 28), month, year) if random.random() < 0.5 else "-"

def generate_attendance_sheet(emp_df, leave_credit_df, leave_availed_df, month: str, year: int):
    attendance_data = []
    month_num = list(calendar.month_name).index(month)
    days_in_month = calendar.monthrange(year, month_num)[1]
    leave_map = get_leave_mapping(leave_availed_df)

    unique_states = emp_df['STATE'].unique().tolist()
    national_holidays, state_holidays = get_random_nh_sh(year, month_num, days_in_month, unique_states, WEEKOFF_DAYS)

    for _, row in emp_df.iterrows():
        record = [
            row['PARENT COMPANY'], row['ASSOCIATE COMPANY'], row['STATE'], row['LOCATION'],
            month, year, row['Employee Code'], row['Employee name']
        ]

        state = row['STATE']
        emp_code = row['Employee Code']
        present_days = 0.0
        for day in range(1, 32):
            if day > days_in_month:
                record.append("-")
                continue

            day_date = date(year, month_num, day)
            day_of_week = day_date.weekday()

            if day_date in national_holidays:
                record.append('NH')
                continue

            if state in state_holidays and day_date in state_holidays[state]:
                record.append('SH')
                continue

            if day_of_week in WEEKOFF_DAYS:
                record.append('WO')
                continue

            if (emp_code, day) in leave_map:
                leave_type = leave_map[(emp_code, day)]
                record.append(leave_type)
                present_days += 0.5 if leave_type == 'HFD' else 0
            else:
                record.append('P')
                present_days += 1

        record.append(present_days)
        attendance_data.append(record)
    return pd.DataFrame(attendance_data, columns=columns_attendance)

def generate_wage_sheet(emp_df, attendance_df, leave_availed_df, month: str, year: int):
    wage_data = []
    month_num = list(calendar.month_name).index(month)
    days_in_month = calendar.monthrange(year, month_num)[1]

    # LOP mapping from leave availed sheet
    lop_map = leave_availed_df[leave_availed_df['LEAVE TYPE'] == 'LOP'].groupby('EMPLOYEE CODE')['NO OF DAYS'].sum().to_dict()

    for _, row in emp_df.iterrows():
        emp_code = row['Employee Code']
        present_days = max(0, days_in_month - lop_map.get(emp_code, 0))
        lop_days = lop_map.get(emp_code, 0)

        # Earnings
        fixed_gross = random.randint(12000, 20000)
        basic = int(fixed_gross * 0.45)
        basic_arrear = random.choice([0, 500])
        hra = int(basic * 0.4)
        hra_arrear = random.choice([0, 300])
        conveyance = random.randint(800, 1500)
        conveyance_arrear = random.choice([0, 200])
        special_allow = random.randint(1000, 2500)
        special_allow_arrear = random.choice([0, 150])
        da = random.randint(800, 2000)
        da_arrear = random.choice([0, 100])
        misc_allow = random.choice([0, 250, 500])
        other_earning = random.choice([0, 400, 800])
        bonus = random.choice([0, 2000, 3000])
        ot_hours = random.randint(0, 10) if random.random() < 0.4 else 0
        ot_amt = ot_hours * 100
        gratuity = int(basic * 0.048)
        medical = 1250
        holiday_pay = random.choice([0, 1000])

        gross_earn = (
            basic + hra + conveyance + special_allow + da + misc_allow +
            other_earning + bonus + ot_amt + gratuity + medical + holiday_pay
        )
        gross_earn_arrear = basic_arrear + hra_arrear + conveyance_arrear + special_allow_arrear + da_arrear

        # Deductions
        pf = int(basic * 0.12)
        pf_arrear = int(basic_arrear * 0.12)
        epf = int(basic * 0.0367)
        pt = random.choice([200, 250])
        pt_arrear = random.choice([0, 50])
        it = random.choice([0, 500])
        esi = int(fixed_gross * 0.0075)
        esi_arrear = int(basic_arrear * 0.0075)
        lwf = 10
        other_ded = random.choice([0, 100])
        fines = random.choice([0, 150])
        insurance = random.choice([0, 500])
        salary_adv = random.choice([0, 2000])

        gross_deduct = pf + pt + it + esi + lwf + other_ded + fines + insurance + salary_adv
        gross_deduct_arrear = pf_arrear + pt_arrear + esi_arrear
        net_pay = gross_earn - gross_deduct
        net_pay_arrear = gross_earn_arrear - gross_deduct_arrear

        # Extra fields
        overtime_date = "{:02d}/{:02d}/{}".format(random.randint(1, 28), month_num, year) if ot_hours > 0 else "-"
        adv_date = "{:02d}/{:02d}/{}".format(random.randint(1, 5), month_num, year) if salary_adv > 0 else "-"
        adv_recovery = random.randint(1, 3) if salary_adv > 0 else 0
        adv_recovery_date = "{:02d}/{:02d}/{}".format(random.randint(20, 28), month_num, year) if salary_adv > 0 else "-"
        fine_date = "{:02d}/{:02d}/{}".format(random.randint(10, 18), month_num, year) if fines > 0 else "-"
        # Randomly assign Relay/Set and Act/Omiss to 50% of employees only
        assign_extra = random.random() < 0.5
        relay_set = random.choice(["Relay", "Set"]) if assign_extra else "-"
        act_omiss = random.choice(["Act", "Omiss"]) if assign_extra else "-"
        fine_nature = random.choice(["Late Submission", "Misconduct"]) if fines > 0 else "-"
        fine_showcause = "{:02d}/{:02d}/{}".format(random.randint(1, 10), month_num, year) if fines > 0 else "-"
        fine_recovery_date = "{:02d}/{:02d}/{}".format(random.randint(25, 28), month_num, year) if fines > 0 else "-"
        loss_amt = random.choice([0, 500, 1000])
        loss_cause_date = "{:02d}/{:02d}/{}".format(random.randint(8, 12), month_num, year) if loss_amt > 0 else "-"
        loss_showcause = "{:02d}/{:02d}/{}".format(random.randint(13, 18), month_num, year) if loss_amt > 0 else "-"
        loss_recovery_inst = random.choice([1, 2]) if loss_amt > 0 else 0
        loss_recovery_done = "{:02d}/{:02d}/{}".format(random.randint(25, days_in_month), month_num, year) if loss_amt > 0 else "-"
        fda_installments = random.choice([1, 2])
        fda_start = "{:02d}/{:02d}/{}".format(random.randint(1, 5), month_num, year)
        fda_end = "{:02d}/{:02d}/{}".format(random.randint(24, days_in_month), month_num, year)
        other_concession = random.choice([0, 1, 2])
        absent_reason = random.choice(["Family Emergency", "Medical", "-", "-"])
        tenure = f"{random.randint(1, 10)} years"
        adv_purpose = random.choice(["Medical", "Family", "Education"]) if salary_adv > 0 else "-"
        adv_deduct_start = adv_date
        adv_deduct_end = adv_recovery_date
        damage_detail = "Laptop damage" if loss_amt > 0 else "-"
        total_contribution = pf + esi
        maternity_from = "{:02d}/{:02d}/{}".format(random.randint(1, 5), month_num, year) if random.random() < 0.15 else "-"
        maternity_child_dob = "{:02d}/{:02d}/{}".format(random.randint(6, 15), month_num, year) if maternity_from != "-" else "-"
        insurance_paid_date = "{:02d}/{:02d}/{}".format(random.randint(10, 20), month_num, year) if random.random() < 0.3 else "-"
        insurance_paid_amt = insurance if insurance_paid_date != "-" else 0
        laid_off = random.choice([0, 1, 2])
        discharge = get_random_date_or_dash(year, month_num)
        proof_ill = get_random_date_or_dash(year, month_num)
        proof_death = get_random_date_or_dash(year, month_num)
        proof_birth = get_random_date_or_dash(year, month_num)
        maternity_paid_date = maternity_child_dob if maternity_from != "-" else "-"
        maternity_advance = random.choice([0, 2000, 3000]) if maternity_from != "-" else 0
        maternity_bonus = random.choice([0, 1500]) if maternity_from != "-" else 0

        record = [
            row['PARENT COMPANY'], row['ASSOCIATE COMPANY'], row['STATE'], row['LOCATION'], row['Branch'],
            month, year, emp_code, present_days, days_in_month, lop_days, fixed_gross, basic, basic_arrear,
            hra, hra_arrear, conveyance, conveyance_arrear, special_allow, special_allow_arrear,
            da, da_arrear, misc_allow, other_earning, bonus, ot_amt, gratuity, medical, holiday_pay,
            gross_earn, gross_earn_arrear, pf, pf_arrear, epf, pt, pt_arrear, it, esi, esi_arrear, lwf,
            other_ded, fines, insurance, salary_adv, gross_deduct, gross_deduct_arrear, net_pay,
            net_pay_arrear, ot_hours, overtime_date, adv_date, adv_recovery, adv_recovery_date,
            fine_nature, relay_set, act_omiss, fine_showcause, fine_recovery_date, loss_amt, loss_cause_date,
            loss_showcause, loss_recovery_inst, loss_recovery_done, fda_installments, fda_start, fda_end,
            other_concession, absent_reason, tenure, adv_purpose, adv_deduct_start, adv_deduct_end,
            damage_detail, total_contribution, maternity_from, maternity_child_dob, insurance_paid_date,
            insurance_paid_amt, laid_off, discharge, proof_ill, proof_death, proof_birth,
            maternity_paid_date, maternity_advance, maternity_bonus
        ]
        wage_data.append(record)

    return pd.DataFrame(wage_data, columns=columns_wage)

def main():
    input_data = {
        "PARENT COMPANY": "InfoZen Private Limited111",
        "ASSOCIATE COMPANY": "InfoZen Private Limited111",
        "STATE_LOCATION": "Andhra Pradesh|Visakhapatnam, Assam|Guwahati, Bihar|Patna, Delhi|Delhi, Chhattisgarh|Raipur, Telangana|Hyderabad, Kerala|Kochi, Punjab|Ludhiana, Maharashtra|Mumbai, West Bengal|Kolkata, Gujarat|Ahmedabad, Haryana|Gurgaon, Madhya Pradesh|Bhopal, Uttar Pradesh|Lucknow, Rajasthan|Bikaner, Goa|Panaji, Odisha|Bhubaneshwar, Chandigarh|Panchkula, Puducherry|Puducherry, Tamil Nadu|Chennai, Karnataka|Bangalore, Jammu and Kashmir|Srinagar, Uttarakhand|DEHRADUN"
    }

    batch_sizes=[450]

    years = [2024]
    company_name = input_data["PARENT COMPANY"].replace(" ", "_")
    os.makedirs(company_name, exist_ok=True)

    # 🔹 Step 1: Generate employees once
    employee_reference_year = min(years)
    emp_df = generate_employee_dataset(input_data, target_year=employee_reference_year)
    emp_df = format_date_columns(emp_df)
    save_in_multiple_batch_sizes(emp_df, "Generated_Employees", batch_sizes, base_output_dir=company_name)

    # 🔹 Step 2: Reuse emp_df across all years/months
    for year in years:
        year_dir = os.path.join(company_name, str(year))
        os.makedirs(year_dir, exist_ok=True)

        for m in range(1, 13):
            month_name = calendar.month_name[m]
            month_dir = os.path.join(year_dir, month_name)
            os.makedirs(month_dir, exist_ok=True)

            leave_credit_df = generate_leave_credit_dataset_realistic(emp_df, month=month_name, year=year)
            leave_availed_df = generate_leave_availed_data(emp_df, month=month_name, year=year)
            attendance_df = generate_attendance_sheet(emp_df, leave_credit_df, leave_availed_df, month=month_name, year=year)
            wage_df = generate_wage_sheet(emp_df, attendance_df, leave_availed_df, month=month_name, year=year)

            leave_credit_df = format_date_columns(leave_credit_df)
            leave_availed_df = format_date_columns(leave_availed_df)

            save_in_multiple_batch_sizes(leave_credit_df, f"Leave_Credit_{month_name}_{year}", batch_sizes, base_output_dir=month_dir)
            save_in_multiple_batch_sizes(leave_availed_df, f"Leave_Availed_{month_name}_{year}", batch_sizes, base_output_dir=month_dir)
            save_in_multiple_batch_sizes(attendance_df, f"Attendance_{month_name}_{year}", batch_sizes, base_output_dir=month_dir)
            save_in_multiple_batch_sizes(wage_df, f"Wage_Sheet_{month_name}_{year}", batch_sizes, base_output_dir=month_dir)

    print("✅ All files generated and saved in split sizes.")

if __name__ == "__main__":
    main()
